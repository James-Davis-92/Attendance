import streamlit as st
import pdfplumber
import pandas as pd
from datetime import datetime, timedelta
from collections import defaultdict
from io import BytesIO
from openpyxl.styles import PatternFill, Alignment
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# Constants
days = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri']

# Actual Google Sheet URL
SHEET_URL = "https://docs.google.com/spreadsheets/d/1Y8nSTQUZ5qWborDOlPxzh9K48Ls2Hpky4vf36Pej7d8/edit?usp=sharing"

# --- Google Sheets functions ---

def get_gsheet_client():
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]

    creds_dict = st.secrets["google"]  # uses your Streamlit Cloud secret
    creds = ServiceAccountCredentials.from_json_keyfile_dict(creds_dict, scope)
    client = gspread.authorize(creds)
    return client

def load_saved_names_from_gs(sheet_url):
    client = get_gsheet_client()
    sheet = client.open_by_url(sheet_url)
    worksheet = sheet.worksheet("names")
    records = worksheet.get_all_records()
    return [(row['Surname'], row['FirstName']) for row in records]

def save_names_to_gs(sheet_url, names_list):
    client = get_gsheet_client()
    sheet = client.open_by_url(sheet_url)
    worksheet = sheet.worksheet("names")
    worksheet.clear()
    worksheet.append_row(["Surname", "FirstName"])
    for surname, first_name in names_list:
        worksheet.append_row([surname, first_name])

# --- PDF + Attendance functions ---

def group_words_to_rows(words, tolerance=3):
    rows, current_row, last_top = [], [], None
    for w in words:
        top = w['top']
        if last_top is None or abs(top - last_top) <= tolerance:
            current_row.append(w)
            last_top = top if last_top is None else (last_top + top) / 2
        else:
            rows.append(current_row)
            current_row, last_top = [w], top
    if current_row:
        rows.append(current_row)
    return rows

def extract_table_from_pdf(pdf_file):
    with pdfplumber.open(pdf_file) as pdf:
        page = pdf.pages[0]
        words = sorted(page.extract_words(), key=lambda w: (w['top'], w['x0']))
        rows = group_words_to_rows(words)
        return [[w['text'] for w in sorted(row, key=lambda w: w['x0'])] for row in rows]

def process_pdf(pdf_file):
    table = extract_table_from_pdf(pdf_file)
    filtered = [row for row in table if len(row) == 12 and row[0] == 'IMSL']
    attendance = {}
    cutoff_time = datetime.strptime("06:01:00", "%H:%M:%S").time()
    for row in filtered:
        surname, first_name, time_str, day_str = row[3].rstrip(','), row[4], row[8], row[6]
        time_obj = datetime.strptime(time_str, "%H:%M:%S").time()
        flag = 'Y' if time_obj < cutoff_time else 'L'
        attendance[(surname, first_name)] = (day_str, flag)
    return attendance

def extract_date_from_filename(filename):
    name, _ = filename.rsplit('.', 1)
    for sep in ['_', '.']:
        parts = name.split(sep)
        if len(parts) >= 3:
            try:
                return datetime(int(parts[2]), int(parts[1]), int(parts[0]))
            except:
                continue
    return None

def style_excel(df):
    with BytesIO() as output:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)
            wb = writer.book
            ws = wb.active

            fill_map = {
                'Y': PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid'),  # Green
                'L': PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid'),  # Yellow
                'A': PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'),  # Red
                'H': PatternFill(start_color='FFFFC0CB', end_color='FFFFC0CB', fill_type='solid'),  # Pink (LightPink)
            }
            center_align = Alignment(horizontal='center', vertical='center')

            # Adjust width for day columns (from column 3 to last column)
            for col_idx in range(3, ws.max_column + 1):
                ws.column_dimensions[chr(64 + col_idx)].width = 20

            for row in range(2, ws.max_row + 1):
                for col_idx in range(3, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if cell.value in fill_map:
                        cell.fill = fill_map[cell.value]
                    cell.alignment = center_align

        return output.getvalue()

# --- Streamlit UI ---

st.title("📋 Attendance Tracker")

# Load saved always-included names from Google Sheets
always_include = load_saved_names_from_gs(SHEET_URL)

# --- Labour List input and save ---
st.subheader("👥 Labour List")

names_str = "\n".join([f"{s}, {f}" for s, f in always_include])

names_input = st.text_area(
    "Enter names (one per line) in the format: Surname, FirstName (Example: Smith, John)",
    value=names_str,
    height=150
)

if st.button("💾 Save names"):
    new_names = []
    for line in names_input.splitlines():
        if ',' in line:
            surname, first_name = map(str.strip, line.split(',', 1))
            if surname and first_name:
                new_names.append((surname, first_name))
    save_names_to_gs(SHEET_URL, new_names)
    st.success("Names saved successfully!")
    always_include = new_names

# --- Weekly attendance uploads ---
uploaded_excel = st.file_uploader("Upload existing weekly attendance Excel file (optional)", type=['xlsx'])

uploaded_pdfs = st.file_uploader(
    "Upload attendance PDF(s) for the week",
    type=["pdf"],
    accept_multiple_files=True
)

if uploaded_pdfs:
    weeks = defaultdict(list)
    for file in uploaded_pdfs:
        date = extract_date_from_filename(file.name)
        if date:
            year, week_num, _ = date.isocalendar()
            week_key = f"{year}-W{week_num:02d}"
            weeks[week_key].append((file, date))
        else:
            st.warning(f"Could not extract date from filename: {file.name}")

    for week_key, files in weeks.items():
        st.subheader(f"📅 Week {week_key}")

        # Load existing Excel data if provided
        if uploaded_excel:
            df_existing = pd.read_excel(uploaded_excel)
            st.write("Loaded existing attendance data:")
            st.dataframe(df_existing)
            # Prepare existing attendance data dict for updating
            all_attendance = defaultdict(lambda: {day: 'A' for day in days})

            # Map current dataframe into all_attendance dictionary
            # Extract base day names (without dates)
            base_days = days
            # Columns from 3rd onwards have date suffixes, map them back to base days:
            col_day_map = {}
            for col in df_existing.columns[2:]:
                # extract the day part from header like 'Mon 01/06/2025'
                base_day = col.split(' ')[0]
                col_day_map[col] = base_day

            for _, row in df_existing.iterrows():
                surname = row['Surname']
                first_name = row['FirstName']
                # Initialize if missing
                if (surname, first_name) not in all_attendance:
                    all_attendance[(surname, first_name)] = {day: 'A' for day in days}
                for col in df_existing.columns[2:]:
                    day = col_day_map[col]
                    all_attendance[(surname, first_name)][day] = row[col]

            # Update attendance from PDFs
            for file, _ in files:
                attendance_for_day = process_pdf(file)
                for (surname, first_name), (day_str, flag) in attendance_for_day.items():
                    if day_str in days:
                        all_attendance[(surname, first_name)][day_str] = flag

            # Ensure always_include names are present
            for name_tuple in always_include:
                if name_tuple not in all_attendance:
                    all_attendance[name_tuple] = {day: 'A' for day in days}

            # Build rows list
            rows = []
            for (surname, first_name), day_flags in all_attendance.items():
                row = [surname, first_name] + [day_flags[day] for day in days]
                rows.append(row)

            df_existing = pd.DataFrame(rows, columns=['Surname', 'FirstName'] + days)
        else:
            # If no existing Excel, build from PDFs only
            all_attendance = defaultdict(lambda: {day: 'A' for day in days})

            for file, _ in files:
                attendance_for_day = process_pdf(file)
                for (surname, first_name), (day_str, flag) in attendance_for_day.items():
                    if day_str in days:
                        all_attendance[(surname, first_name)][day_str] = flag

            for name_tuple in always_include:
                if name_tuple not in all_attendance:
                    all_attendance[name_tuple] = {day: 'A' for day in days}

            rows = []
            for (surname, first_name), day_flags in all_attendance.items():
                row = [surname, first_name] + [day_flags[day] for day in days]
                rows.append(row)

            df_existing = pd.DataFrame(rows, columns=['Surname', 'FirstName'] + days)
        
        # Add date suffixes to days in columns
        year, week_num = map(int, week_key.split('-W'))
        monday_date = datetime.strptime(f"{year} {week_num} 1", "%G %V %u")
        day_headers = [f"{day} {(monday_date + timedelta(days=i)).strftime('%d/%m/%Y')}" for i, day in enumerate(days)]
        df_existing.columns = ['Surname', 'FirstName'] + day_headers
        
        st.dataframe(df_existing)

        excel_bytes = style_excel(df_existing)
        st.download_button(
            label=f"⬇️ Download updated Excel for week {week_key}",
            data=excel_bytes,
            file_name=f"attendance_{week_key}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
else:
    st.info("Upload PDFs to process weekly attendance.")

